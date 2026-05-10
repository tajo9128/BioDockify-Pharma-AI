from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape
import xml.etree.ElementTree as ET

from plugins._office.helpers import document_store, pptx_writer


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
P_NS = "http://schemas.openxmlformats.org/presentationml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XML_NS = "http://www.w3.org/XML/1998/namespace"
ODF_OFFICE_NS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
ODF_TEXT_NS = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
ODF_TABLE_NS = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
ODF_DRAW_NS = "urn:oasis:names:tc:opendocument:xmlns:drawing:1.0"
ODF_PRESENTATION_NS = "urn:oasis:names:tc:opendocument:xmlns:presentation:1.0"
ODS_DIRECT_EDIT_ROW_LIMIT = 10000
ODS_DIRECT_EDIT_COLUMN_LIMIT = 1024

for prefix, namespace in {
    "w": W_NS,
    "a": A_NS,
    "p": P_NS,
    "r": R_NS,
    "office": ODF_OFFICE_NS,
    "text": ODF_TEXT_NS,
    "table": ODF_TABLE_NS,
    "draw": ODF_DRAW_NS,
    "presentation": ODF_PRESENTATION_NS,
}.items():
    ET.register_namespace(prefix, namespace)


def qn(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def read_artifact(doc: dict[str, Any], max_chars: int = 12000) -> dict[str, Any]:
    """Extract compact editable content from an Office artifact."""
    path = Path(doc["path"])
    ext = str(doc["extension"]).lower()
    if ext == "md":
        content = _read_markdown(path)
    elif ext == "odt":
        content = _read_odt(path)
    elif ext == "ods":
        content = _read_ods(path)
    elif ext == "odp":
        content = _read_odp(path)
    elif ext == "docx":
        content = _read_docx(path)
    elif ext == "xlsx":
        content = _read_xlsx(path)
    elif ext == "pptx":
        content = _read_pptx(path)
    else:
        raise ValueError(f"Unsupported document format: {ext}")

    return _trim_payload(content, max_chars=max_chars)


def edit_artifact(
    doc: dict[str, Any],
    operation: str = "",
    content: str = "",
    find: str = "",
    replace: str = "",
    sheet: str = "",
    cells: Any = None,
    rows: Any = None,
    chart: Any = None,
    slides: Any = None,
    **kwargs: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Apply a direct saved edit to an Office artifact and return updated metadata."""
    path = Path(doc["path"])
    ext = str(doc["extension"]).lower()
    op = normalize_operation(operation, content=content, find=find, cells=cells, rows=rows, chart=chart, slides=slides)
    before = path.read_bytes()

    invalidate_sessions = bool(kwargs.pop("invalidate_sessions", False))
    if ext == "md":
        updated, details = _edit_markdown(before, op, content=content, find=find, replace=replace, **kwargs)
    elif ext == "odt":
        updated, details = _edit_odt(before, op, content=content, find=find, replace=replace, **kwargs)
    elif ext == "ods":
        updated, details = _edit_ods(before, op, content=content, find=find, replace=replace, sheet=sheet, cells=cells, rows=rows, **kwargs)
    elif ext == "odp":
        updated, details = _edit_odp(before, op, content=content, find=find, replace=replace, slides=slides, **kwargs)
    elif ext == "docx":
        updated, details = _edit_docx(before, op, content=content, find=find, replace=replace, **kwargs)
    elif ext == "xlsx":
        updated, details = _edit_xlsx(path, op, content=content, find=find, replace=replace, sheet=sheet, cells=cells, rows=rows, chart=chart, **kwargs)
    elif ext == "pptx":
        updated, details = _edit_pptx(before, op, content=content, find=find, replace=replace, slides=slides, **kwargs)
    else:
        raise ValueError(f"Direct edit is not available for .{ext}.")

    changed = updated != before
    updated_doc = (
        document_store.replace_document_bytes(
            doc["file_id"],
            updated,
            actor="document_artifact:edit",
            invalidate_sessions=invalidate_sessions,
        )
        if changed
        else doc
    )
    if changed:
        _refresh_open_editor_sessions(updated_doc["file_id"])
    preview = read_artifact(updated_doc, max_chars=int(kwargs.get("preview_chars") or 4000))
    payload = {
        "ok": True,
        "action": "edit",
        "operation": op,
        "changed": changed,
        **details,
        "preview": preview,
    }
    return updated_doc, payload


def _refresh_open_editor_sessions(file_id: str) -> None:
    try:
        from plugins._office.helpers import markdown_sessions

        markdown_sessions.get_manager().refresh_document(file_id)
    except Exception:
        # Direct artifact edits should never fail just because no canvas is open.
        pass
    try:
        from plugins._office.helpers import libreoffice_desktop

        libreoffice_desktop.get_manager().refresh_document(file_id)
    except Exception:
        pass


def normalize_operation(
    operation: str,
    *,
    content: str = "",
    find: str = "",
    cells: Any = None,
    rows: Any = None,
    chart: Any = None,
    slides: Any = None,
) -> str:
    op = str(operation or "").strip().lower().replace("-", "_")
    aliases = {
        "patch": "replace_text" if find else "set_text",
        "update": "replace_text" if find else "set_text",
        "replace": "replace_text",
        "append": "append_text",
        "prepend": "prepend_text",
        "write": "set_text",
        "set": "set_text",
        "set_content": "set_text",
        "set_sheet": "set_rows",
        "write_sheet": "set_rows",
        "add_rows": "append_rows",
        "add_chart": "create_chart",
        "chart": "create_chart",
        "insert_chart": "create_chart",
        "set_chart": "create_chart",
        "add_slide": "append_slide",
        "set_deck": "set_slides",
    }
    op = aliases.get(op, op)
    if op:
        return op
    if cells:
        return "set_cells"
    if rows:
        return "append_rows"
    if chart:
        return "create_chart"
    if slides:
        return "set_slides"
    if find:
        return "replace_text"
    if content:
        return "set_text"
    raise ValueError("operation is required")


def _read_markdown(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    lines = [line for line in text.splitlines() if line.strip()]
    headings = [line.lstrip("#").strip() for line in lines if line.lstrip().startswith("#")]
    return {
        "kind": "document",
        "format": "markdown",
        "line_count": len(text.splitlines()),
        "headings": headings[:40],
        "text": text,
    }


def _read_odt(path: Path) -> dict[str, Any]:
    root = _odf_content_root(path)
    paragraphs = _odf_text_lines(root)
    headings = [
        "".join(node.itertext()).strip()
        for node in root.iter(qn(ODF_TEXT_NS, "h"))
        if "".join(node.itertext()).strip()
    ]
    return {
        "kind": "document",
        "format": "odt",
        "paragraph_count": len(paragraphs),
        "headings": headings[:40],
        "text": "\n".join(paragraphs),
        "paragraphs": paragraphs[:80],
    }


def _read_ods(path: Path) -> dict[str, Any]:
    sheets = _ods_sheets_from_bytes(
        path.read_bytes(),
        max_rows=ODS_DIRECT_EDIT_ROW_LIMIT,
        max_cols=ODS_DIRECT_EDIT_COLUMN_LIMIT,
    )
    return {
        "kind": "spreadsheet",
        "format": "ods",
        "sheet_count": len(sheets),
        "sheets": [
            {
                "name": sheet["name"],
                "max_row": len(sheet["rows"]),
                "max_column": max((len(row) for row in sheet["rows"]), default=0),
                "chart_count": 0,
                "charts": [],
                "preview_rows": sheet["rows"][:80],
            }
            for sheet in sheets[:8]
        ],
    }


def _read_odp(path: Path) -> dict[str, Any]:
    slides = _odp_text_slides(path.read_bytes())
    return {
        "kind": "presentation",
        "format": "odp",
        "slide_count": len(slides),
        "slides": [
            {
                "index": index + 1,
                "title": slide.get("title", ""),
                "lines": [slide.get("title", ""), *slide.get("bullets", [])],
            }
            for index, slide in enumerate(slides[:40])
        ],
    }


def _read_docx(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    paragraphs = []
    for paragraph in root.iter(qn(W_NS, "p")):
        text = "".join(node.text or "" for node in paragraph.iter(qn(W_NS, "t")))
        if text.strip():
            paragraphs.append(text)
    return {
        "kind": "document",
        "paragraph_count": len(paragraphs),
        "text": "\n".join(paragraphs),
        "paragraphs": paragraphs[:80],
    }


def _read_xlsx(path: Path) -> dict[str, Any]:
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(path, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets[:8]:
        rows = []
        max_row = min(worksheet.max_row or 0, 80)
        max_col = min(worksheet.max_column or 0, 30)
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            values = ["" if value is None else value for value in row]
            if any(str(value).strip() for value in values):
                rows.append(values)
        charts = [_chart_summary(chart) for chart in getattr(worksheet, "_charts", [])[:20]]
        sheets.append({
            "name": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "chart_count": len(getattr(worksheet, "_charts", [])),
            "charts": charts,
            "preview_rows": rows,
        })
    return {
        "kind": "spreadsheet",
        "sheet_count": len(workbook.worksheets),
        "sheets": sheets,
    }


def _read_pptx(path: Path) -> dict[str, Any]:
    slides = []
    with zipfile.ZipFile(path) as archive:
        for name in _slide_names(archive):
            root = ET.fromstring(archive.read(name))
            lines = []
            for paragraph in root.iter(qn(A_NS, "p")):
                text = "".join(node.text or "" for node in paragraph.iter(qn(A_NS, "t"))).strip()
                if text:
                    lines.append(text)
            slides.append({
                "index": len(slides) + 1,
                "title": lines[0] if lines else "",
                "lines": lines,
            })
    return {
        "kind": "presentation",
        "slide_count": len(slides),
        "slides": slides[:40],
    }


def _edit_markdown(before: bytes, op: str, *, content: str = "", find: str = "", replace: str = "", **kwargs: Any) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "append_text", "prepend_text", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported Markdown operation: {op}")

    text = before.decode("utf-8", errors="replace")
    if op == "set_text":
        updated = content
        details = {"lines_written": len(content.splitlines())}
    elif op == "append_text":
        separator = "" if not text or text.endswith("\n") else "\n"
        updated = f"{text}{separator}{content}"
        details = {"lines_appended": len(content.splitlines())}
    elif op == "prepend_text":
        separator = "" if not text or content.endswith("\n") else "\n"
        updated = f"{content}{separator}{text}"
        details = {"lines_prepended": len(content.splitlines())}
    else:
        if not find:
            raise ValueError("find is required for replace_text")
        replacement = "" if op == "delete_text" else replace
        count_limit = _int_or_none(kwargs.get("count"))
        updated, count = _replace_limited(text, find, replacement, count_limit)
        details = {"replacements": count}
    return updated.encode("utf-8"), details


def _edit_odt(before: bytes, op: str, *, content: str = "", find: str = "", replace: str = "", **kwargs: Any) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "append_text", "prepend_text", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported ODT operation: {op}")

    paragraphs = _odf_text_lines(ET.fromstring(_zip_member(before, "content.xml")))
    if op == "set_text":
        lines = _text_lines(content)
        return document_store.odt_bytes_from_paragraphs(lines), {"paragraphs_written": len(lines)}
    if op == "append_text":
        lines = [*paragraphs, *_text_lines(content)]
        return document_store.odt_bytes_from_paragraphs(lines), {"paragraphs_written": len(lines)}
    if op == "prepend_text":
        lines = [*_text_lines(content), *paragraphs]
        return document_store.odt_bytes_from_paragraphs(lines), {"paragraphs_written": len(lines)}

    if not find:
        raise ValueError("find is required for replace_text")
    replacement = "" if op == "delete_text" else replace
    joined, count = _replace_limited("\n".join(paragraphs), find, replacement, _int_or_none(kwargs.get("count")))
    if count == 0:
        return before, {"replacements": count}
    return document_store.odt_bytes_from_paragraphs(joined.splitlines()), {"replacements": count}


def _edit_docx(before: bytes, op: str, *, content: str = "", find: str = "", replace: str = "", **kwargs: Any) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "append_text", "prepend_text", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported DOCX operation: {op}")

    with zipfile.ZipFile(io.BytesIO(before)) as archive:
        files = {info.filename: archive.read(info.filename) for info in archive.infolist()}
    root = ET.fromstring(files["word/document.xml"])

    if op == "replace_text" or op == "delete_text":
        if not find:
            raise ValueError("find is required for replace_text")
        replacement = "" if op == "delete_text" else replace
        count = _replace_text_in_paragraphs(
            root,
            paragraph_tag=qn(W_NS, "p"),
            text_tag=qn(W_NS, "t"),
            set_text=_set_word_paragraph_text,
            find=find,
            replacement=replacement,
            limit=_int_or_none(kwargs.get("count")),
        )
        details = {"replacements": count}
        if count == 0:
            return before, details
    else:
        lines = _text_lines(content)
        body = root.find(f".//{qn(W_NS, 'body')}")
        if body is None:
            raise ValueError("DOCX document body not found")
        paragraphs = [_word_paragraph(line) for line in lines]
        if op == "set_text":
            sect_pr = [child for child in list(body) if child.tag == qn(W_NS, "sectPr")]
            for child in list(body):
                body.remove(child)
            for paragraph in paragraphs:
                body.append(paragraph)
            for child in sect_pr:
                body.append(child)
        elif op == "append_text":
            insert_at = len(body)
            for idx, child in enumerate(list(body)):
                if child.tag == qn(W_NS, "sectPr"):
                    insert_at = idx
                    break
            for paragraph in reversed(paragraphs):
                body.insert(insert_at, paragraph)
        elif op == "prepend_text":
            for paragraph in reversed(paragraphs):
                body.insert(0, paragraph)
        details = {"paragraphs_written": len(paragraphs)}

    files["word/document.xml"] = _xml_bytes(root)
    return _zip_from_existing(files), details


def _edit_ods(
    before: bytes,
    op: str,
    *,
    content: str = "",
    find: str = "",
    replace: str = "",
    sheet: str = "",
    cells: Any = None,
    rows: Any = None,
    **kwargs: Any,
) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "set_rows", "append_text", "append_rows", "set_cells", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported ODS operation: {op}")

    sheets = _ods_sheets_from_bytes(
        before,
        max_rows=ODS_DIRECT_EDIT_ROW_LIMIT,
        max_cols=ODS_DIRECT_EDIT_COLUMN_LIMIT,
        strict_limits=True,
    )
    if not sheets:
        sheets = [{"name": "Sheet1", "rows": []}]
    worksheet = _ods_sheet(sheets, sheet)
    details: dict[str, Any] = {"sheet": worksheet["name"]}

    if op in {"set_text", "set_rows"}:
        parsed_rows = _normalize_rows(rows if rows is not None else content)
        worksheet["rows"] = parsed_rows
        details["rows_written"] = len(parsed_rows)
    elif op in {"append_text", "append_rows"}:
        parsed_rows = _normalize_rows(rows if rows is not None else content)
        worksheet["rows"].extend(parsed_rows)
        details["rows_appended"] = len(parsed_rows)
        details["start_row"] = max(len(worksheet["rows"]) - len(parsed_rows) + 1, 1)
    elif op == "set_cells":
        assignments = _normalize_cells(cells, default_sheet=worksheet["name"])
        for sheet_name, cell, value in assignments:
            target = _ods_sheet(sheets, sheet_name)
            row_idx, col_idx = _cell_indices(cell)
            _set_matrix_value(target["rows"], row_idx, col_idx, value)
        details["cells_written"] = len(assignments)
    else:
        if not find:
            raise ValueError("find is required for replace_text")
        replacement = "" if op == "delete_text" else replace
        count = 0
        limit = _int_or_none(kwargs.get("count"))
        for item in sheets:
            for row_idx, row in enumerate(item["rows"]):
                for col_idx, value in enumerate(row):
                    if not isinstance(value, str) or find not in value:
                        continue
                    remaining = None if limit is None else max(limit - count, 0)
                    if remaining == 0:
                        break
                    row[col_idx], replaced = _replace_limited(value, find, replacement, remaining)
                    count += replaced
                if limit is not None and count >= limit:
                    break
            if limit is not None and count >= limit:
                break
        details["replacements"] = count
        if count == 0:
            return before, details

    return document_store.ods_bytes_from_sheets(sheets), details


def _edit_xlsx(
    path: Path,
    op: str,
    *,
    content: str = "",
    find: str = "",
    replace: str = "",
    sheet: str = "",
    cells: Any = None,
    rows: Any = None,
    chart: Any = None,
    **kwargs: Any,
) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "set_rows", "append_text", "append_rows", "set_cells", "replace_text", "delete_text", "create_chart"}:
        raise ValueError(f"Unsupported XLSX operation: {op}")
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(path)
    worksheet = _worksheet(workbook, sheet)

    details: dict[str, Any] = {"sheet": worksheet.title}
    if op in {"set_text", "set_rows"}:
        parsed_rows = _normalize_rows(rows if rows is not None else content)
        _clear_worksheet(worksheet)
        _write_rows(worksheet, parsed_rows, start_row=1)
        details["rows_written"] = len(parsed_rows)
    elif op in {"append_text", "append_rows"}:
        parsed_rows = _normalize_rows(rows if rows is not None else content)
        start_row = max((worksheet.max_row or 0) + 1, 1)
        _write_rows(worksheet, parsed_rows, start_row=start_row)
        details["rows_appended"] = len(parsed_rows)
        details["start_row"] = start_row
    elif op == "set_cells":
        assignments = _normalize_cells(cells, default_sheet=worksheet.title)
        for sheet_name, cell, value in assignments:
            target = _worksheet(workbook, sheet_name)
            target[cell] = value
        details["cells_written"] = len(assignments)
    elif op in {"replace_text", "delete_text"}:
        if not find:
            raise ValueError("find is required for replace_text")
        replacement = "" if op == "delete_text" else replace
        count = 0
        limit = _int_or_none(kwargs.get("count"))
        for target in workbook.worksheets:
            for row in target.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or find not in cell.value:
                        continue
                    remaining = None if limit is None else max(limit - count, 0)
                    if remaining == 0:
                        break
                    cell.value, replaced = _replace_limited(cell.value, find, replacement, remaining)
                    count += replaced
                if limit is not None and count >= limit:
                    break
            if limit is not None and count >= limit:
                break
        details["replacements"] = count
        if count == 0:
            return path.read_bytes(), details
    elif op == "create_chart":
        chart_details = []
        for chart_spec in _normalize_chart_specs(chart, kwargs):
            chart_details.append(_create_xlsx_chart(workbook, worksheet, chart_spec))
        details["charts_created"] = len(chart_details)
        details["charts"] = chart_details

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), details


_CHART_SPEC_KEYS = {
    "anchor",
    "categories",
    "chart_type",
    "close",
    "data_range",
    "fields",
    "from_rows",
    "height",
    "high",
    "include_headers",
    "labels",
    "legend",
    "low",
    "open",
    "position",
    "replace_existing",
    "series",
    "sheet",
    "style",
    "title",
    "titles_from_data",
    "type",
    "values",
    "width",
    "x_axis_title",
    "xvalues",
    "y_axis_title",
    "yvalues",
}

_CHART_TYPE_ALIASES = {
    "area": "area",
    "bar": "bar",
    "candlestick": "stock",
    "col": "column",
    "column": "column",
    "columns": "column",
    "line": "line",
    "ohlc": "stock",
    "pie": "pie",
    "scatter": "scatter",
    "stock": "stock",
}


def _normalize_chart_specs(chart: Any, kwargs: dict[str, Any]) -> list[dict[str, Any]]:
    parsed = _parse_chart_value(chart)
    if isinstance(parsed, list):
        if not parsed:
            raise ValueError("chart list must include at least one chart spec")
        return [_normalize_chart_spec(item, {}) for item in parsed]
    if parsed is None:
        parsed = {}
    if not isinstance(parsed, dict):
        raise ValueError("chart must be an object, JSON object, or list of chart objects")
    return [_normalize_chart_spec(parsed, kwargs)]


def _parse_chart_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if stripped.startswith("{") or stripped.startswith("["):
            return json.loads(stripped)
        return {"type": stripped}
    return value


def _normalize_chart_spec(value: Any, kwargs: dict[str, Any]) -> dict[str, Any]:
    if isinstance(value, str):
        value = _parse_chart_value(value)
    if value is None:
        value = {}
    if not isinstance(value, dict):
        raise ValueError("each chart spec must be an object")

    spec = dict(value)
    explicit_include_headers = "include_headers" in spec or "titles_from_data" in spec
    for key in _CHART_SPEC_KEYS:
        if key in kwargs and kwargs[key] not in (None, ""):
            spec[key] = kwargs[key]
            if key in {"include_headers", "titles_from_data"}:
                explicit_include_headers = True

    explicit_type = bool(spec.get("type") or spec.get("chart_type"))
    chart_type = str(spec.get("type") or spec.get("chart_type") or "").strip().lower().replace("-", "_")
    if chart_type:
        chart_type = _CHART_TYPE_ALIASES.get(chart_type, chart_type)
    spec["type"] = chart_type
    spec["_explicit_type"] = explicit_type
    spec["position"] = str(spec.get("position") or spec.get("anchor") or "H2")
    spec["include_headers"] = _bool_value(spec.get("include_headers", spec.get("titles_from_data")), default=True)
    spec["_include_headers_explicit"] = explicit_include_headers
    spec["from_rows"] = _bool_value(spec.get("from_rows"), default=False)
    spec["replace_existing"] = _bool_value(spec.get("replace_existing"), default=False)
    spec["width"] = _float_or_default(spec.get("width"), 18.0)
    spec["height"] = _float_or_default(spec.get("height"), 10.0)
    return spec


def _create_xlsx_chart(workbook: Any, default_worksheet: Any, spec: dict[str, Any]) -> dict[str, Any]:
    openpyxl = _require_openpyxl()
    worksheet = _worksheet(workbook, str(spec.get("sheet") or default_worksheet.title))
    chart_type = spec["type"] or _infer_default_chart_type(worksheet)
    if chart_type not in _CHART_TYPE_ALIASES.values():
        raise ValueError(f"Unsupported XLSX chart type: {chart_type}")

    if spec["replace_existing"]:
        charts_removed = len(getattr(worksheet, "_charts", []))
        worksheet._charts = []
    else:
        charts_removed = 0

    if chart_type == "stock":
        chart, data_range, categories = _stock_chart(openpyxl, workbook, worksheet, spec)
    elif chart_type == "scatter":
        chart, data_range, categories = _scatter_chart(openpyxl, workbook, worksheet, spec)
    else:
        chart, data_range, categories = _standard_chart(openpyxl, workbook, worksheet, spec, chart_type)

    _apply_chart_options(chart, spec)
    worksheet.add_chart(chart, spec["position"])
    return {
        "type": chart_type,
        "title": str(spec.get("title") or ""),
        "sheet": worksheet.title,
        "position": spec["position"],
        "data_range": data_range,
        "categories": categories,
        "series_count": len(getattr(chart, "series", [])),
        "charts_removed": charts_removed,
    }


def _standard_chart(openpyxl: Any, workbook: Any, worksheet: Any, spec: dict[str, Any], chart_type: str) -> tuple[Any, str, str]:
    chart_classes = {
        "area": openpyxl.chart.AreaChart,
        "bar": openpyxl.chart.BarChart,
        "column": openpyxl.chart.BarChart,
        "line": openpyxl.chart.LineChart,
        "pie": openpyxl.chart.PieChart,
    }
    chart = chart_classes[chart_type]()
    if chart_type == "bar":
        chart.type = "bar"
    elif chart_type == "column":
        chart.type = "col"

    include_headers = bool(spec["include_headers"])
    categories = str(spec.get("categories") or spec.get("labels") or "")
    if spec.get("series"):
        data_range = _add_explicit_series(openpyxl, chart, workbook, worksheet, spec, validate_numeric=True)
    else:
        range_value = spec.get("values") or spec.get("data_range") or _default_data_range(worksheet, chart_type, include_headers)
        include_headers = _include_headers_for_range(spec, range_value)
        data_ref, data_sheet, bounds, data_range = _reference_from_range(openpyxl, workbook, worksheet, range_value)
        _validate_numeric_series(data_sheet, bounds, include_headers=include_headers, label=data_range)
        chart.add_data(data_ref, titles_from_data=include_headers, from_rows=bool(spec["from_rows"]))

    if not categories:
        categories = _default_category_range(worksheet, include_headers=include_headers)
    if categories:
        categories_ref, _, _, categories = _reference_from_range(openpyxl, workbook, worksheet, categories)
        chart.set_categories(categories_ref)
    return chart, data_range, categories


def _stock_chart(openpyxl: Any, workbook: Any, worksheet: Any, spec: dict[str, Any]) -> tuple[Any, str, str]:
    chart = openpyxl.chart.StockChart()
    field_ranges = _stock_field_ranges(spec)

    if field_ranges:
        data_labels = []
        for label in ("open", "high", "low", "close"):
            include_headers = _include_headers_for_range(spec, field_ranges[label])
            series_ref, data_sheet, bounds, label_range = _reference_from_range(openpyxl, workbook, worksheet, field_ranges[label])
            _validate_numeric_series(data_sheet, bounds, include_headers=include_headers, label=label)
            chart.series.append(openpyxl.chart.Series(series_ref, title_from_data=include_headers))
            data_labels.append(label_range)
        data_range = ", ".join(data_labels)
    elif spec.get("series"):
        include_headers = bool(spec["include_headers"])
        data_range = _add_explicit_series(openpyxl, chart, workbook, worksheet, spec, expected_count=4, validate_numeric=True)
    else:
        include_headers = bool(spec["include_headers"])
        range_value = spec.get("data_range") or _default_data_range(worksheet, "stock", include_headers)
        include_headers = _include_headers_for_range(spec, range_value)
        _, data_sheet, bounds, range_label = _reference_from_range(openpyxl, workbook, worksheet, range_value)
        min_col, min_row, max_col, max_row = bounds
        columns = list(range(min_col, max_col + 1))
        if len(columns) > 4 and _looks_like_category_header(data_sheet.cell(row=min_row, column=min_col).value):
            columns = columns[1:5]
        else:
            columns = columns[:4]
        if len(columns) != 4:
            raise ValueError("stock charts require exactly four Open, High, Low, Close data series")
        _validate_stock_headers(data_sheet, columns, min_row, include_headers=include_headers)
        for column in columns:
            _validate_numeric_series(data_sheet, (column, min_row, column, max_row), include_headers=include_headers, label=data_sheet.cell(row=min_row, column=column).value or _column_letter(column))
            series_ref = openpyxl.chart.Reference(data_sheet, min_col=column, min_row=min_row, max_col=column, max_row=max_row)
            chart.series.append(openpyxl.chart.Series(series_ref, title_from_data=include_headers))
        data_range = range_label

    categories = str(spec.get("categories") or spec.get("labels") or _default_category_range(worksheet, include_headers=bool(spec["include_headers"])))
    if categories:
        categories_ref, _, _, categories = _reference_from_range(openpyxl, workbook, worksheet, categories)
        chart.set_categories(categories_ref)
    chart.hiLowLines = openpyxl.chart.axis.ChartLines()
    chart.upDownBars = openpyxl.chart.updown_bars.UpDownBars()
    return chart, data_range, categories


def _scatter_chart(openpyxl: Any, workbook: Any, worksheet: Any, spec: dict[str, Any]) -> tuple[Any, str, str]:
    chart = openpyxl.chart.ScatterChart()
    include_headers = bool(spec["include_headers"])
    categories = str(spec.get("xvalues") or spec.get("categories") or _default_category_range(worksheet, include_headers=include_headers))
    x_ref, x_sheet, x_bounds, categories = _reference_from_range(openpyxl, workbook, worksheet, categories)
    if include_headers and x_bounds[1] == 1 and x_bounds[3] > 1:
        x_ref = openpyxl.chart.Reference(x_sheet, min_col=x_bounds[0], min_row=2, max_col=x_bounds[2], max_row=x_bounds[3])

    data_ranges = []
    series_items = _series_items(spec)
    if series_items:
        for item in series_items:
            values_ref, title, data_range = _series_values_reference(openpyxl, workbook, worksheet, item, include_headers=include_headers, validate_numeric=True)
            xvalues = item.get("xvalues") or item.get("x") or item.get("categories")
            if xvalues:
                item_x_ref, item_x_sheet, item_x_bounds, _ = _reference_from_range(openpyxl, workbook, worksheet, xvalues)
                if include_headers and item_x_bounds[1] == 1 and item_x_bounds[3] > 1:
                    item_x_ref = openpyxl.chart.Reference(item_x_sheet, min_col=item_x_bounds[0], min_row=2, max_col=item_x_bounds[2], max_row=item_x_bounds[3])
            else:
                item_x_ref = x_ref
            chart.series.append(openpyxl.chart.Series(values_ref, xvalues=item_x_ref, title=title))
            data_ranges.append(data_range)
    else:
        range_value = spec.get("yvalues") or spec.get("values") or spec.get("data_range") or _default_data_range(worksheet, "scatter", include_headers)
        _, data_sheet, bounds, data_range = _reference_from_range(openpyxl, workbook, worksheet, range_value)
        min_col, min_row, max_col, max_row = bounds
        first_row = min_row + 1 if include_headers and min_row == 1 and max_row > 1 else min_row
        for column in range(min_col, max_col + 1):
            title = data_sheet.cell(row=min_row, column=column).value if first_row > min_row else None
            _validate_numeric_series(data_sheet, (column, min_row, column, max_row), include_headers=include_headers, label=title or _column_letter(column))
            y_ref = openpyxl.chart.Reference(data_sheet, min_col=column, min_row=first_row, max_col=column, max_row=max_row)
            chart.series.append(openpyxl.chart.Series(y_ref, xvalues=x_ref, title=str(title) if title is not None else None))
    return chart, ", ".join(data_ranges) if data_ranges else data_range, categories


def _add_explicit_series(
    openpyxl: Any,
    chart: Any,
    workbook: Any,
    worksheet: Any,
    spec: dict[str, Any],
    expected_count: int | None = None,
    validate_numeric: bool = False,
) -> str:
    include_headers = bool(spec["include_headers"])
    ranges = []
    for item in _series_items(spec):
        values_ref, title, label = _series_values_reference(openpyxl, workbook, worksheet, item, include_headers=include_headers, validate_numeric=validate_numeric)
        if title:
            chart.series.append(openpyxl.chart.Series(values_ref, title=title))
        else:
            chart.series.append(openpyxl.chart.Series(values_ref, title_from_data=include_headers))
        ranges.append(label)
    if expected_count is not None and len(ranges) != expected_count:
        raise ValueError(f"chart requires exactly {expected_count} series")
    return ", ".join(ranges)


def _series_items(spec: dict[str, Any]) -> list[dict[str, Any]]:
    raw = spec.get("series") or []
    if isinstance(raw, str):
        raw = json.loads(raw)
    if not isinstance(raw, list):
        raise ValueError("chart series must be a list")
    items = []
    for item in raw:
        if isinstance(item, str):
            items.append({"values": item})
        elif isinstance(item, dict):
            items.append(item)
        else:
            raise ValueError("chart series entries must be objects or range strings")
    return items


def _series_values_reference(
    openpyxl: Any,
    workbook: Any,
    worksheet: Any,
    item: dict[str, Any],
    *,
    include_headers: bool,
    validate_numeric: bool = False,
) -> tuple[Any, str | None, str]:
    values = item.get("values") or item.get("range") or item.get("yvalues") or item.get("y")
    if not values:
        raise ValueError("chart series entries require values or range")
    ref, data_sheet, bounds, label = _reference_from_range(openpyxl, workbook, worksheet, values)
    title = item.get("title") or item.get("name")
    min_col, min_row, max_col, max_row = bounds
    if include_headers and min_row == 1 and max_col == min_col and max_row > min_row:
        title = title if title is not None else data_sheet.cell(row=min_row, column=min_col).value
        ref = openpyxl.chart.Reference(data_sheet, min_col=min_col, min_row=min_row + 1, max_col=max_col, max_row=max_row)
    if validate_numeric:
        _validate_numeric_series(data_sheet, bounds, include_headers=include_headers, label=title or label)
    return ref, str(title) if title is not None else None, label


def _stock_field_ranges(spec: dict[str, Any]) -> dict[str, Any]:
    fields = spec.get("fields") or {}
    if isinstance(fields, str):
        fields = json.loads(fields)
    if not isinstance(fields, dict):
        raise ValueError("stock chart fields must be an object")
    result = {}
    for label in ("open", "high", "low", "close"):
        value = spec.get(label) or fields.get(label)
        if value:
            result[label] = value
    if result and set(result) != {"open", "high", "low", "close"}:
        raise ValueError("stock chart fields must include open, high, low, and close")
    return result


def _reference_from_range(openpyxl: Any, workbook: Any, default_worksheet: Any, value: Any) -> tuple[Any, Any, tuple[int, int, int, int], str]:
    sheet_name, cell_range = _split_range_ref(value, default_worksheet.title)
    worksheet = _worksheet(workbook, sheet_name)
    min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(cell_range)
    reference = openpyxl.chart.Reference(worksheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    return reference, worksheet, (min_col, min_row, max_col, max_row), _range_label(openpyxl, worksheet.title, min_col, min_row, max_col, max_row)


def _split_range_ref(value: Any, default_sheet: str) -> tuple[str, str]:
    if isinstance(value, dict):
        sheet = str(value.get("sheet") or default_sheet)
        min_cell = value.get("range") or value.get("ref")
        if min_cell:
            return _split_range_ref(str(min_cell), sheet)
        min_col = value.get("min_col")
        min_row = value.get("min_row")
        max_col = value.get("max_col", min_col)
        max_row = value.get("max_row", min_row)
        if min_col is None or min_row is None:
            raise ValueError("range objects require range/ref or min_col and min_row")
        return sheet, f"{_cell_ref(min_col, min_row)}:{_cell_ref(max_col, max_row)}"
    ref = str(value or "").strip()
    if not ref:
        raise ValueError("chart range is required")
    if "!" not in ref:
        return default_sheet, ref
    sheet, cell_range = ref.rsplit("!", 1)
    return sheet.strip().strip("'") or default_sheet, cell_range


def _range_label(openpyxl: Any, sheet_title: str, min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    start = f"{openpyxl.utils.cell.get_column_letter(min_col)}{min_row}"
    end = f"{openpyxl.utils.cell.get_column_letter(max_col)}{max_row}"
    sheet = sheet_title if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", sheet_title) else f"'{sheet_title}'"
    return f"{sheet}!{start}:{end}"


def _include_headers_for_range(spec: dict[str, Any], range_value: Any) -> bool:
    include_headers = bool(spec["include_headers"])
    if spec.get("_include_headers_explicit"):
        return include_headers
    try:
        _, cell_range = _split_range_ref(range_value, "")
        _, min_row, _, _ = __import__("openpyxl").utils.cell.range_boundaries(cell_range)
    except Exception:
        return include_headers
    return include_headers and min_row == 1


def _validate_stock_headers(data_sheet: Any, columns: list[int], min_row: int, *, include_headers: bool) -> None:
    if not include_headers or min_row != 1:
        return
    expected = ["open", "high", "low", "close"]
    found = [str(data_sheet.cell(row=min_row, column=column).value or "").strip().lower() for column in columns]
    if found != expected:
        raise ValueError(f"stock charts require Open, High, Low, Close columns in order; found {found}")


def _validate_numeric_series(data_sheet: Any, bounds: tuple[int, int, int, int], *, include_headers: bool, label: Any) -> None:
    min_col, min_row, max_col, max_row = bounds
    start_row = min_row + 1 if include_headers and min_row == 1 and max_row > min_row else min_row
    values = [
        data_sheet.cell(row=row, column=column).value
        for column in range(min_col, max_col + 1)
        for row in range(start_row, max_row + 1)
    ]
    numeric_count = sum(1 for value in values if isinstance(value, (int, float)) and not isinstance(value, bool))
    if numeric_count == 0:
        name = str(label or _range_label(__import__("openpyxl"), data_sheet.title, min_col, min_row, max_col, max_row))
        raise ValueError(f"chart series '{name}' has no numeric data")


def _default_category_range(worksheet: Any, *, include_headers: bool) -> str:
    if (worksheet.max_column or 0) < 2 or (worksheet.max_row or 0) < 2:
        return ""
    first_row_is_header = _looks_like_category_header(worksheet.cell(row=1, column=1).value)
    return f"A{2 if include_headers or first_row_is_header else 1}:A{worksheet.max_row}"


def _default_data_range(worksheet: Any, chart_type: str, include_headers: bool) -> str:
    max_row = worksheet.max_row or 1
    max_col = worksheet.max_column or 1
    if chart_type == "stock":
        if max_col < 5 or max_row < 2:
            raise ValueError("stock charts need Date, Open, High, Low, Close columns or explicit ranges")
        return f"B{1 if include_headers else 2}:E{max_row}"
    if chart_type == "pie" and max_col >= 2:
        return f"B{1 if include_headers else 2}:B{max_row}"
    start_col = 2 if max_col >= 2 else 1
    return f"{_column_letter(start_col)}{1 if include_headers else 2}:{_column_letter(max_col)}{max_row}"


def _cell_ref(column: Any, row: Any) -> str:
    return f"{_column_letter(column)}{int(row)}"


def _column_letter(column: Any) -> str:
    if isinstance(column, str) and column.isalpha():
        return column.upper()
    return __import__("openpyxl").utils.cell.get_column_letter(int(column))


def _infer_default_chart_type(worksheet: Any) -> str:
    headers = [str(worksheet.cell(row=1, column=column).value or "").strip().lower() for column in range(1, (worksheet.max_column or 0) + 1)]
    if {"open", "high", "low", "close"}.issubset(set(headers)):
        return "stock"
    return "line"


def _looks_like_category_header(value: Any) -> bool:
    return str(value or "").strip().lower() in {"date", "time", "category", "label", "month", "year"}


def _apply_chart_options(chart: Any, spec: dict[str, Any]) -> None:
    if spec.get("title"):
        chart.title = str(spec["title"])
    if spec.get("style") not in (None, ""):
        chart.style = int(spec["style"])
    chart.width = spec["width"]
    chart.height = spec["height"]
    if _bool_value(spec.get("legend"), default=True) is False:
        chart.legend = None
    if spec.get("x_axis_title") and hasattr(chart, "x_axis"):
        chart.x_axis.title = str(spec["x_axis_title"])
    if spec.get("y_axis_title") and hasattr(chart, "y_axis"):
        chart.y_axis.title = str(spec["y_axis_title"])


def _chart_summary(chart: Any) -> dict[str, Any]:
    return {
        "type": _chart_kind(chart),
        "title": _chart_title(chart),
        "anchor": _chart_anchor(chart),
        "series_count": len(getattr(chart, "series", [])),
    }


def _chart_kind(chart: Any) -> str:
    name = chart.__class__.__name__.replace("Chart", "").lower()
    return {"bar": "bar_or_column", "stock": "stock"}.get(name, name)


def _chart_title(chart: Any) -> str:
    title = getattr(chart, "title", None)
    if title is None or isinstance(title, str):
        return title or ""
    try:
        paragraphs = title.tx.rich.p
        parts = []
        for paragraph in paragraphs:
            for run in paragraph.r:
                if run.t:
                    parts.append(run.t)
        return "".join(parts)
    except Exception:
        return ""


def _chart_anchor(chart: Any) -> str:
    openpyxl = _require_openpyxl()
    anchor = getattr(chart, "anchor", "")
    if isinstance(anchor, str):
        return anchor
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return ""
    return f"{openpyxl.utils.cell.get_column_letter(marker.col + 1)}{marker.row + 1}"


def _bool_value(value: Any, default: bool = False) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() not in {"0", "false", "no", "off", "none"}


def _float_or_default(value: Any, default: float) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _edit_pptx(before: bytes, op: str, *, content: str = "", find: str = "", replace: str = "", slides: Any = None, **kwargs: Any) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "set_slides", "append_text", "append_slide", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported PPTX operation: {op}")

    if op in {"set_text", "set_slides"}:
        parsed_slides = _normalize_slides(slides if slides is not None else content)
        return _pptx_from_slides(parsed_slides), {"slides_written": len(parsed_slides)}

    if op in {"append_text", "append_slide"}:
        existing = _pptx_text_slides(before)
        existing.extend(_normalize_slides(slides if slides is not None else content))
        return _pptx_from_slides(existing), {"slides_written": len(existing)}

    with zipfile.ZipFile(io.BytesIO(before)) as archive:
        files = {info.filename: archive.read(info.filename) for info in archive.infolist()}
    if not find:
        raise ValueError("find is required for replace_text")
    replacement = "" if op == "delete_text" else replace
    count = 0
    limit = _int_or_none(kwargs.get("count"))
    for name in sorted([name for name in files if name.startswith("ppt/slides/slide") and name.endswith(".xml")], key=_natural_key):
        root = ET.fromstring(files[name])
        count += _replace_text_in_paragraphs(
            root,
            paragraph_tag=qn(A_NS, "p"),
            text_tag=qn(A_NS, "t"),
            set_text=_set_drawing_paragraph_text,
            find=find,
            replacement=replacement,
            limit=None if limit is None else max(limit - count, 0),
        )
        files[name] = _xml_bytes(root)
        if limit is not None and count >= limit:
            break
    if count == 0:
        return before, {"replacements": count}
    return _zip_from_existing(files), {"replacements": count}


def _edit_odp(before: bytes, op: str, *, content: str = "", find: str = "", replace: str = "", slides: Any = None, **kwargs: Any) -> tuple[bytes, dict[str, Any]]:
    if op not in {"set_text", "set_slides", "append_text", "append_slide", "replace_text", "delete_text"}:
        raise ValueError(f"Unsupported ODP operation: {op}")

    if op in {"set_text", "set_slides"}:
        parsed_slides = _normalize_slides(slides if slides is not None else content)
        return document_store.odp_bytes_from_slides(parsed_slides), {"slides_written": len(parsed_slides)}

    existing = _odp_text_slides(before)
    if op in {"append_text", "append_slide"}:
        existing.extend(_normalize_slides(slides if slides is not None else content))
        return document_store.odp_bytes_from_slides(existing), {"slides_written": len(existing)}

    if not find:
        raise ValueError("find is required for replace_text")
    replacement = "" if op == "delete_text" else replace
    count = 0
    limit = _int_or_none(kwargs.get("count"))
    for slide in existing:
        title, title_count = _replace_limited(
            str(slide.get("title") or ""),
            find,
            replacement,
            None if limit is None else max(limit - count, 0),
        )
        if title_count:
            slide["title"] = title
            count += title_count
        if limit is not None and count >= limit:
            break
        bullets = []
        for bullet in slide.get("bullets") or []:
            updated, replaced = _replace_limited(
                str(bullet),
                find,
                replacement,
                None if limit is None else max(limit - count, 0),
            )
            bullets.append(updated)
            count += replaced
            if limit is not None and count >= limit:
                bullets.extend(slide.get("bullets", [])[len(bullets):])
                break
        slide["bullets"] = bullets
        if limit is not None and count >= limit:
            break
    if count == 0:
        return before, {"replacements": count}
    return document_store.odp_bytes_from_slides(existing), {"replacements": count}


def _replace_text_in_paragraphs(
    root: ET.Element,
    *,
    paragraph_tag: str,
    text_tag: str,
    set_text: Any,
    find: str,
    replacement: str,
    limit: int | None,
) -> int:
    count = 0
    for paragraph in root.iter(paragraph_tag):
        texts = list(paragraph.iter(text_tag))
        if not texts:
            continue
        current = "".join(node.text or "" for node in texts)
        if find not in current:
            continue
        remaining = None if limit is None else max(limit - count, 0)
        if remaining == 0:
            break
        updated, replaced = _replace_limited(current, find, replacement, remaining)
        if replaced:
            set_text(paragraph, updated)
            count += replaced
    return count


def _replace_limited(value: str, find: str, replacement: str, limit: int | None) -> tuple[str, int]:
    if limit is None:
        return value.replace(find, replacement), value.count(find)
    return value.replace(find, replacement, limit), min(value.count(find), limit)


def _set_word_paragraph_text(paragraph: ET.Element, text: str) -> None:
    keep = [child for child in list(paragraph) if child.tag == qn(W_NS, "pPr")]
    for child in list(paragraph):
        paragraph.remove(child)
    for child in keep:
        paragraph.append(child)
    paragraph.append(_word_run(text))


def _word_paragraph(text: str) -> ET.Element:
    paragraph = ET.Element(qn(W_NS, "p"))
    paragraph.append(_word_run(text))
    return paragraph


def _word_run(text: str) -> ET.Element:
    run = ET.Element(qn(W_NS, "r"))
    text_node = ET.SubElement(run, qn(W_NS, "t"))
    if text.startswith(" ") or text.endswith(" "):
        text_node.set(qn(XML_NS, "space"), "preserve")
    text_node.text = text
    return run


def _set_drawing_paragraph_text(paragraph: ET.Element, text: str) -> None:
    keep = [child for child in list(paragraph) if child.tag == qn(A_NS, "pPr")]
    for child in list(paragraph):
        paragraph.remove(child)
    for child in keep:
        paragraph.append(child)
    run = ET.SubElement(paragraph, qn(A_NS, "r"))
    text_node = ET.SubElement(run, qn(A_NS, "t"))
    text_node.text = text


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for spreadsheet edits") from exc
    return openpyxl


def _worksheet(workbook: Any, sheet: str = "") -> Any:
    if sheet:
        if sheet not in workbook.sheetnames:
            return workbook.create_sheet(sheet)
        return workbook[sheet]
    return workbook.active


def _clear_worksheet(worksheet: Any) -> None:
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)


def _write_rows(worksheet: Any, rows: list[list[Any]], start_row: int) -> None:
    for row_offset, row in enumerate(rows):
        for col_offset, value in enumerate(row):
            worksheet.cell(row=start_row + row_offset, column=1 + col_offset, value=_cell_value(value))


def _normalize_rows(value: Any) -> list[list[Any]]:
    if value is None:
        return []
    if isinstance(value, list):
        rows = value
    elif isinstance(value, str):
        rows = _rows_from_text(value)
    else:
        rows = [[value]]
    normalized = []
    for row in rows:
        if isinstance(row, (list, tuple)):
            normalized.append([_cell_value(value) for value in row])
        else:
            normalized.append([_cell_value(row)])
    return normalized


def _normalize_cells(cells: Any, default_sheet: str) -> list[tuple[str, str, Any]]:
    if isinstance(cells, str):
        parsed = json.loads(cells)
    else:
        parsed = cells
    if not parsed:
        raise ValueError("cells is required for set_cells")

    result: list[tuple[str, str, Any]] = []
    if isinstance(parsed, dict):
        for ref, value in parsed.items():
            sheet, cell = _split_cell_ref(str(ref), default_sheet)
            result.append((sheet, cell, _cell_value(value)))
    elif isinstance(parsed, list):
        for item in parsed:
            if not isinstance(item, dict):
                raise ValueError("cells list entries must be objects")
            ref = str(item.get("cell") or item.get("ref") or "")
            sheet = str(item.get("sheet") or default_sheet)
            if "!" in ref:
                sheet, ref = _split_cell_ref(ref, default_sheet)
            if not ref:
                raise ValueError("cell is required for each cells entry")
            result.append((sheet, ref, _cell_value(item.get("value"))))
    else:
        raise ValueError("cells must be an object or list")
    return result


def _split_cell_ref(ref: str, default_sheet: str) -> tuple[str, str]:
    if "!" not in ref:
        return default_sheet, ref
    sheet, cell = ref.split("!", 1)
    return sheet.strip("'") or default_sheet, cell


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _rows_from_text(content: str) -> list[list[str]]:
    text = str(content or "").strip("\n")
    if not text.strip():
        return []
    lines = [line for line in text.splitlines() if line.strip()]
    markdown_rows = _markdown_table_rows(lines)
    if markdown_rows:
        return markdown_rows

    delimiter = "\t" if any("\t" in line for line in lines) else ("," if any("," in line for line in lines) else None)
    if delimiter:
        return [row for row in csv.reader(io.StringIO("\n".join(lines)), delimiter=delimiter)]
    return [[line] for line in lines]


def _markdown_table_rows(lines: list[str]) -> list[list[str]]:
    table_lines = [line.strip() for line in lines if line.strip().startswith("|") and line.strip().endswith("|")]
    if len(table_lines) < 2:
        return []
    rows = []
    for line in table_lines:
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if all(re.fullmatch(r":?-{3,}:?", cell or "") for cell in cells):
            continue
        rows.append(cells)
    return rows


def _zip_member(data: bytes, name: str) -> bytes:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        return archive.read(name)


def _odf_content_root(path: Path) -> ET.Element:
    with zipfile.ZipFile(path) as archive:
        return ET.fromstring(archive.read("content.xml"))


def _odf_text_lines(root: ET.Element) -> list[str]:
    lines = []
    for node in root.iter():
        if node.tag not in {qn(ODF_TEXT_NS, "h"), qn(ODF_TEXT_NS, "p")}:
            continue
        text = "".join(node.itertext()).strip()
        if text:
            lines.append(text)
    return lines


def _ods_sheets_from_bytes(
    data: bytes,
    *,
    max_rows: int | None = None,
    max_cols: int | None = None,
    strict_limits: bool = False,
) -> list[dict[str, Any]]:
    root = ET.fromstring(_zip_member(data, "content.xml"))
    sheets = []
    for index, table in enumerate(root.iter(qn(ODF_TABLE_NS, "table")), start=1):
        name = table.get(qn(ODF_TABLE_NS, "name")) or table.get("name") or f"Sheet{index}"
        rows = []
        for row in table:
            if row.tag != qn(ODF_TABLE_NS, "table-row"):
                continue
            values = _ods_row_values(row, max_cols=max_cols, strict_limits=strict_limits)
            repeat_rows = _repeat_count(row.get(qn(ODF_TABLE_NS, "number-rows-repeated")))
            append_count = repeat_rows
            if max_rows is not None:
                remaining = max(max_rows - len(rows), 0)
                append_count = min(repeat_rows, remaining)
                if strict_limits and repeat_rows > append_count and _row_has_content(values):
                    raise ValueError(f"ODS direct editing is limited to {max_rows} populated rows per sheet.")
                if remaining == 0:
                    if not strict_limits:
                        break
                    continue
            for _ in range(append_count):
                rows.append(values.copy())
            if max_rows is not None and len(rows) >= max_rows and not strict_limits:
                break
        sheets.append({"name": name, "rows": _trim_blank_edges(rows)})
    return sheets


def _ods_row_values(
    row: ET.Element,
    *,
    max_cols: int | None = None,
    strict_limits: bool = False,
) -> list[Any]:
    values = []
    for cell in row:
        if cell.tag not in {qn(ODF_TABLE_NS, "table-cell"), qn(ODF_TABLE_NS, "covered-table-cell")}:
            continue
        value = _ods_cell_value(cell)
        repeat = _repeat_count(cell.get(qn(ODF_TABLE_NS, "number-columns-repeated")))
        append_count = repeat
        if max_cols is not None:
            remaining = max(max_cols - len(values), 0)
            append_count = min(repeat, remaining)
            if strict_limits and repeat > append_count and _cell_has_content(value):
                raise ValueError(f"ODS direct editing is limited to {max_cols} populated columns per sheet.")
            if remaining == 0:
                if not strict_limits:
                    break
                continue
        for _ in range(append_count):
            values.append(value)
        if max_cols is not None and len(values) >= max_cols and not strict_limits:
            break
    return values


def _ods_cell_value(cell: ET.Element) -> Any:
    value_type = str(cell.get(qn(ODF_OFFICE_NS, "value-type")) or "").lower()
    if value_type in {"float", "currency", "percentage"}:
        raw = cell.get(qn(ODF_OFFICE_NS, "value"))
        if raw not in (None, ""):
            try:
                number = float(raw)
                return int(number) if number.is_integer() else number
            except ValueError:
                pass
    if value_type == "boolean":
        raw = str(cell.get(qn(ODF_OFFICE_NS, "boolean-value")) or "").lower()
        if raw in {"true", "false"}:
            return raw == "true"
    text = "\n".join("".join(node.itertext()).strip() for node in cell.iter(qn(ODF_TEXT_NS, "p")))
    return text.strip()


def _repeat_count(value: Any) -> int:
    try:
        count = int(value or 1)
    except (TypeError, ValueError):
        count = 1
    return max(1, count)


def _trim_blank_edges(rows: list[list[Any]]) -> list[list[Any]]:
    trimmed = []
    for row in rows:
        next_row = list(row)
        while next_row and not _cell_has_content(next_row[-1]):
            next_row.pop()
        trimmed.append(next_row)
    while trimmed and not _row_has_content(trimmed[-1]):
        trimmed.pop()
    return trimmed


def _cell_has_content(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _row_has_content(row: list[Any]) -> bool:
    return any(_cell_has_content(value) for value in row)


def _ods_sheet(sheets: list[dict[str, Any]], name: str = "") -> dict[str, Any]:
    normalized = str(name or "").strip()
    if normalized:
        for sheet in sheets:
            if str(sheet["name"]).casefold() == normalized.casefold():
                return sheet
        sheet = {"name": normalized, "rows": []}
        sheets.append(sheet)
        return sheet
    return sheets[0]


def _cell_indices(cell: str) -> tuple[int, int]:
    match = re.fullmatch(r"\$?([A-Za-z]{1,4})\$?([1-9][0-9]*)", str(cell or "").strip())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell}")
    col = 0
    for char in match.group(1).upper():
        col = col * 26 + (ord(char) - 64)
    return int(match.group(2)), col


def _set_matrix_value(rows: list[list[Any]], row_idx: int, col_idx: int, value: Any) -> None:
    while len(rows) < row_idx:
        rows.append([])
    row = rows[row_idx - 1]
    while len(row) < col_idx:
        row.append("")
    row[col_idx - 1] = value


def _odp_text_slides(data: bytes) -> list[dict[str, Any]]:
    root = ET.fromstring(_zip_member(data, "content.xml"))
    slides = []
    for page in root.iter(qn(ODF_DRAW_NS, "page")):
        lines = []
        for node in page.iter():
            if node.tag not in {qn(ODF_TEXT_NS, "h"), qn(ODF_TEXT_NS, "p")}:
                continue
            text = "".join(node.itertext()).strip()
            if text:
                lines.append(text)
        if lines:
            slides.append({"title": lines[0], "bullets": lines[1:]})
    return slides


def _pptx_text_slides(data: bytes) -> list[dict[str, Any]]:
    slides = []
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        for name in _slide_names(archive):
            root = ET.fromstring(archive.read(name))
            lines = []
            for paragraph in root.iter(qn(A_NS, "p")):
                text = "".join(node.text or "" for node in paragraph.iter(qn(A_NS, "t"))).strip()
                if text:
                    lines.append(text)
            if lines:
                slides.append({"title": lines[0], "bullets": lines[1:]})
    return slides


def _normalize_slides(value: Any) -> list[dict[str, Any]]:
    if value is None:
        return []
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        if stripped.startswith("[") or stripped.startswith("{"):
            return _normalize_slides(json.loads(stripped))
        chunks = re.split(r"(?m)^\s*---+\s*$", stripped)
        result = []
        for chunk in chunks:
            lines = [line.strip(" -\t") for line in chunk.splitlines() if line.strip()]
            if not lines:
                continue
            result.append({"title": lines[0], "bullets": lines[1:]})
        return result
    if isinstance(value, dict):
        return [_slide_from_mapping(value)]
    if isinstance(value, list):
        result = []
        for item in value:
            if isinstance(item, dict):
                result.append(_slide_from_mapping(item))
            elif isinstance(item, str):
                result.extend(_normalize_slides(item))
            elif isinstance(item, (list, tuple)):
                lines = [str(part) for part in item if str(part).strip()]
                if lines:
                    result.append({"title": lines[0], "bullets": lines[1:]})
            else:
                result.append({"title": str(item), "bullets": []})
        return result
    return [{"title": str(value), "bullets": []}]


def _slide_from_mapping(value: dict[str, Any]) -> dict[str, Any]:
    title = str(value.get("title") or value.get("heading") or "Slide")
    bullets = value.get("bullets")
    if bullets is None:
        body = value.get("body") or value.get("content") or ""
        bullets = [line.strip(" -\t") for line in str(body).splitlines() if line.strip()]
    elif isinstance(bullets, str):
        bullets = [line.strip(" -\t") for line in bullets.splitlines() if line.strip()]
    else:
        bullets = [str(item) for item in bullets]
    return {"title": title, "bullets": bullets}


def _pptx_from_slides(slides: list[dict[str, Any]]) -> bytes:
    return pptx_writer.pptx_from_slides(slides)


def _pptx_content_types(count: int) -> str:
    overrides = [
        '<Override PartName="/ppt/presentation.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>'
    ]
    for index in range(1, count + 1):
        overrides.append(
            f'<Override PartName="/ppt/slides/slide{index}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<Types xmlns="{CT_NS}">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        + "".join(overrides)
        + "</Types>"
    )


def _pptx_presentation_rels(count: int) -> str:
    rels = []
    for index in range(1, count + 1):
        rels.append(
            f'<Relationship Id="rId{index}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" '
            f'Target="slides/slide{index}.xml"/>'
        )
    return '<?xml version="1.0" encoding="UTF-8"?>' + f'<Relationships xmlns="{REL_NS}">' + "".join(rels) + "</Relationships>"


def _pptx_presentation_xml(count: int) -> str:
    slide_ids = "".join(f'<p:sldId id="{255 + index}" r:id="rId{index}"/>' for index in range(1, count + 1))
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<p:presentation xmlns:p="{P_NS}" xmlns:r="{R_NS}">'
        f"<p:sldIdLst>{slide_ids}</p:sldIdLst>"
        '<p:sldSz cx="9144000" cy="5143500"/>'
        "</p:presentation>"
    )


def _pptx_slide_xml(slide: dict[str, Any]) -> str:
    title = str(slide.get("title") or "Slide")
    bullets = [str(item) for item in slide.get("bullets") or []]
    paragraphs = [title, *bullets]
    text = "".join(f"<a:p><a:r><a:t>{escape(item)}</a:t></a:r></a:p>" for item in paragraphs)
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f'<p:sld xmlns:a="{A_NS}" xmlns:p="{P_NS}">'
        "<p:cSld><p:spTree>"
        '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
        "<p:grpSpPr/>"
        '<p:sp><p:nvSpPr><p:cNvPr id="2" name="Content"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>'
        f"<p:txBody><a:bodyPr/><a:lstStyle/>{text}</p:txBody>"
        "</p:sp>"
        "</p:spTree></p:cSld>"
        "</p:sld>"
    )


def _slide_names(archive: zipfile.ZipFile) -> list[str]:
    return sorted(
        [name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")],
        key=_natural_key,
    )


def _natural_key(value: str) -> list[Any]:
    return [int(part) if part.isdigit() else part for part in re.split(r"(\d+)", value)]


def _text_lines(content: str) -> list[str]:
    lines = [line.rstrip() for line in str(content or "").splitlines()]
    return lines or [""]


def _int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def _xml_bytes(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _zip_from_existing(files: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, data in files.items():
            archive.writestr(name, data)
    return buffer.getvalue()


def _zip_map(files: dict[str, str | bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, value in files.items():
            archive.writestr(name, value.encode("utf-8") if isinstance(value, str) else value)
    return buffer.getvalue()


def _trim_payload(payload: dict[str, Any], max_chars: int) -> dict[str, Any]:
    text = json.dumps(payload, ensure_ascii=False, default=str)
    if len(text) <= max_chars:
        return payload
    trimmed = dict(payload)
    if "paragraphs" in trimmed:
        trimmed["paragraphs"] = trimmed["paragraphs"][:20]
    if "sheets" in trimmed:
        trimmed["sheets"] = [
            {**sheet, "preview_rows": sheet.get("preview_rows", [])[:20]}
            for sheet in trimmed["sheets"][:4]
        ]
    if "slides" in trimmed:
        trimmed["slides"] = trimmed["slides"][:12]
    if "text" in trimmed and isinstance(trimmed["text"], str):
        trimmed["text"] = trimmed["text"][:max_chars] + "\n... [truncated]"
    trimmed["truncated"] = True
    return trimmed
